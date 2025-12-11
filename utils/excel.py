from io import BytesIO
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# =============================================================================
# 1. CARGA FLEXIBLE DE INVENTARIO BASE
# =============================================================================
def load_inventory_excel(file):
    df = pd.read_excel(file)

    columnas = [
        "Código del Material",
        "Texto breve de material",
        "Unidad de medida base",
        "Ubicación",
        "Libre utilización",
    ]

    for c in columnas:
        if c not in df.columns:
            raise Exception(f"❌ Falta columna: {c}")

    df["Código del Material"] = df["Código del Material"].astype(str).str.strip()
    df["Ubicación"] = df["Ubicación"].astype(str).str.strip()

    return df


# =============================================================================
# 2. CARGA DE INVENTARIO 2D (REQUERIDO POR warehouse2d_routes)
# =============================================================================
def load_warehouse2d_excel(file):
    df = pd.read_excel(file)

    columnas_requeridas = [
        "Código del Material",
        "Texto breve de material",
        "Unidad de medida base",
        "Ubicación",
        "Stock máximo",
        "Consumo mes actual",
        "Libre utilización",
        "Tamaño de lote mínimo",
    ]

    for c in columnas_requeridas:
        if c not in df.columns:
            raise Exception(f"❌ Falta columna obligatoria en mapa 2D: {c}")

    df["Código del Material"] = df["Código del Material"].astype(str).str.strip()
    df["Ubicación"] = df["Ubicación"].astype(str).str.strip()

    return df


# =============================================================================
# 3. ORDENAR UBICACIONES (E001, E015, E120...)
# =============================================================================
def sort_location_advanced(loc):
    try:
        if isinstance(loc, str) and loc.startswith("E"):
            return int("".join([x for x in loc if x.isdigit()]))
        return 999999
    except:
        return 999999


# =============================================================================
# 4. GENERAR EXCEL DE DISCREPANCIAS (SIN CORRUPCIÓN)
# =============================================================================
def generate_discrepancies_excel(df):

    output = BytesIO()

    # Si viene sin datos
    if df is None or df.empty:
        wb = Workbook()
        ws = wb.active
        ws.title = "Discrepancias"
        ws.append(["SIN DATOS"])
        wb.save(output)
        output.seek(0)
        return output

    # Normalizar todo a texto
    df = df.astype(str)

    wb = Workbook()
    ws = wb.active
    ws.title = "Discrepancias"

    # Cabeceras
    ws.append(list(df.columns))

    # Filas
    for _, row in df.iterrows():
        ws.append(row.tolist())

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    center = Alignment(horizontal="center")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # Bordes
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = center

    ws.auto_filter.ref = ws.dimensions

    wb.save(output)
    output.seek(0)
    return output
